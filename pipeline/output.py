"""
output.py — Structured Output Generation (Phase 2)
────────────────────────────────────────────────────
Writes results to:
  - Excel (.xlsx)  — 12-column schema with flag distribution + confidence summary
  - JSON           — full audit record per figure
  - Console log    — processing metrics matching Phase 2 spec
"""

import json
from pathlib import Path
from datetime import datetime
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed — Excel output disabled")


# ── Colour palette ────────────────────────────────────────────────────────────
HEADER_FILL       = "2D5FA0"   # dark blue
HEADER_FONT       = "FFFFFF"
FLAG_OK_FILL      = "E8F5E9"   # green
FLAG_INSUFF_FILL  = "FFF9C4"   # yellow
FLAG_MISS_FILL    = "FFE0B2"   # orange
FLAG_ERROR_FILL   = "FFEBEE"   # red
FLAG_SKIP_FILL    = "F5F5F5"   # grey
ALT_ROW_FILL      = "F5F8FF"   # subtle alternating


# ── Phase 2 — 12-column schema ────────────────────────────────────────────────
COLUMNS = [
    ("source_file",        18, "File"),
    ("page_num",            6, "Page"),
    ("fig_id",             28, "Figure ID"),
    ("fig_type",           16, "Type"),
    ("image_filename",     30, "Image Path"),
    ("caption",            48, "Caption"),
    ("caption_word_count", 10, "Caption Words"),
    ("caption_flag",       20, "Caption Flag"),
    ("image_status",       16, "Image Status"),
    ("existing_alt",       36, "Existing Alt"),
    ("confidence",         12, "Confidence"),
    ("final_flag",         24, "Final Flag"),
    ("alt_text",           56, "Generated Alt"),
    ("status",             14, "Status"),
]


def _header_style(cell):
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.font = Font(bold=True, color=HEADER_FONT, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _flag_fill(final_flag: str) -> str:
    if final_flag == "OK":
        return FLAG_OK_FILL
    if "INSUFFICIENT" in final_flag:
        return FLAG_INSUFF_FILL
    if "MISSING" in final_flag or "NOT_RESOLVED" in final_flag:
        return FLAG_MISS_FILL
    if "ERROR" in final_flag:
        return FLAG_ERROR_FILL
    if "SKIP" in final_flag or "PRESENT" in final_flag:
        return FLAG_SKIP_FILL
    return "FFFFFF"


def write_excel(figures: list, output_path: str) -> str:
    if not HAS_OPENPYXL:
        print("[SKIP] openpyxl not available")
        return ""

    output_path = Path(output_path)
    try:
        wb = openpyxl.Workbook()
        _write_alt_text_sheet(wb, figures)
        _write_summary_sheet(wb, figures)
        wb.save(str(output_path))
        print(f"📊 Excel saved: {output_path}")
        return str(output_path)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = output_path.with_name(f"{output_path.stem}_{ts}{output_path.suffix}")
        wb.save(str(fallback))
        print(f"[WARN] File locked — saved as: {fallback.name}")
        return str(fallback)


def _write_alt_text_sheet(wb, figures: list):
    ws = wb.active
    ws.title = "Alt Text"
    ws.freeze_panes = "A2"

    for col_idx, (key, width, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, fig in enumerate(figures, start=2):
        fill_hex = _flag_fill(getattr(fig, "final_flag", ""))
        if row_idx % 2 == 0 and fill_hex == FLAG_OK_FILL:
            fill_hex = ALT_ROW_FILL
        fill = PatternFill("solid", fgColor=fill_hex)

        values = [
            getattr(fig, "source_file", getattr(fig, "pdf_name", "")),
            fig.page_num,
            fig.fig_id,
            fig.fig_type,
            fig.image_filename,
            fig.caption,
            getattr(fig, "caption_word_count", ""),
            getattr(fig, "caption_flag", ""),
            getattr(fig, "image_status", ""),
            getattr(fig, "existing_alt", ""),
            getattr(fig, "confidence", ""),
            getattr(fig, "final_flag", ""),
            fig.alt_text,
            fig.status,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.font = Font(size=10)
        ws.row_dimensions[row_idx].height = 40


def _write_summary_sheet(wb, figures: list):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12

    current_row = 1

    def _section(title, counts):
        nonlocal current_row
        h1 = ws.cell(row=current_row, column=1, value=title)
        h2 = ws.cell(row=current_row, column=2, value="Count")
        _header_style(h1)
        _header_style(h2)
        current_row += 1
        for label, count in sorted(counts.items(), key=lambda x: -x[1]):
            ws.cell(row=current_row, column=1, value=label).font = Font(size=10)
            ws.cell(row=current_row, column=2, value=count).font = Font(size=10)
            current_row += 1
        total = ws.cell(row=current_row, column=1, value="TOTAL")
        total.font = Font(bold=True, size=10)
        ws.cell(row=current_row, column=2, value=sum(counts.values())).font = Font(bold=True, size=10)
        current_row += 3

    _section("Flag Distribution",   Counter(getattr(f, "final_flag", "unknown") for f in figures))
    _section("Caption Flag",        Counter(getattr(f, "caption_flag", "unknown") for f in figures))
    _section("Confidence",          Counter(str(getattr(f, "confidence", "N/A")) for f in figures))
    _section("Figure Type",         Counter(f.fig_type for f in figures))
    _section("Generation Status",   Counter(f.status for f in figures))
    _section("Source Type",         Counter(getattr(f, "source_type", "pdf") for f in figures))


def write_json(figures: list, output_path: str) -> str:
    output_path = Path(output_path)
    records = []
    for fig in figures:
        records.append({
            "source_file":        getattr(fig, "source_file", getattr(fig, "pdf_name", "")),
            "source_type":        getattr(fig, "source_type", "pdf"),
            "page":               fig.page_num,
            "fig_id":             fig.fig_id,
            "fig_type":           fig.fig_type,
            "image":              fig.image_filename,
            "caption":            fig.caption,
            "caption_word_count": getattr(fig, "caption_word_count", None),
            "caption_flag":       getattr(fig, "caption_flag", None),
            "image_status":       getattr(fig, "image_status", None),
            "existing_alt":       getattr(fig, "existing_alt", ""),
            "confidence":         getattr(fig, "confidence", None),
            "final_flag":         getattr(fig, "final_flag", None),
            "alt_text":           fig.alt_text,
            "status":             fig.status,
            "processing_status":  getattr(fig, "processing_status", "PROCESSED"),
            "xml_fig_element_id": getattr(fig, "xml_fig_element_id", ""),
        })
    output_path.write_text(json.dumps(records, indent=2, ensure_ascii=False))
    print(f"📄 JSON saved: {output_path}")
    return str(output_path)


def print_metrics(figures: list, source_name: str = ""):
    """Print Phase 2 processing summary matching the spec's sample output."""
    total = len(figures)
    flag_counts     = Counter(getattr(f, "final_flag", "unknown") for f in figures)
    conf_counts     = Counter(str(getattr(f, "confidence", "N/A")) for f in figures)
    generated       = sum(1 for f in figures if f.status == "done")
    already_present = sum(1 for f in figures if f.status == "skipped_existing_alt")
    skipped         = sum(1 for f in figures if f.status == "skipped")
    errors          = sum(1 for f in figures if "error" in f.status.lower())

    label = f" — {source_name}" if source_name else ""
    print(f"""
========== PROCESSING SUMMARY{label} ==========
Total Figures : {total}

---------- VALIDATION ----------
OK                        : {flag_counts.get('OK', 0)}
IMAGE_MISSING             : {flag_counts.get('IMAGE_MISSING', 0)}
IMAGE_NOT_RESOLVED        : {flag_counts.get('IMAGE_NOT_RESOLVED', 0)}
CAPTION_MISSING           : {flag_counts.get('CAPTION_MISSING', 0)}
CAPTION_INSUFFICIENT      : {flag_counts.get('CAPTION_INSUFFICIENT', 0)}
IMAGE_AND_CAPTION_MISSING : {flag_counts.get('IMAGE_AND_CAPTION_MISSING', 0)}
ALT_ALREADY_PRESENT       : {flag_counts.get('ALT_ALREADY_PRESENT', 0)}

---------- GENERATION ----------
Alt Text Generated        : {generated}
Already Present (kept)    : {already_present}
Skipped                   : {skipped}
Errors                    : {errors}

---------- CONFIDENCE ----------
HIGH                      : {conf_counts.get('HIGH', 0)}
MEDIUM                    : {conf_counts.get('MEDIUM', 0)}
LOW                       : {conf_counts.get('LOW', 0)}
N/A                       : {conf_counts.get('None', 0) + conf_counts.get('N/A', 0)}
==========================================""")